#!/usr/bin/env python3
"""
lit_review_matrix.py — Generator literature review matrix (CSV/XLSX).

Buat template dan isi literature review matrix untuk organize referensi
dan identify research gap.

Usage CLI:
  python3 lit_review_matrix.py create --output matrix.csv
  python3 lit_review_matrix.py create --output matrix.xlsx
  python3 lit_review_matrix.py add --file matrix.csv --author "Smith" --year 2023 \\
    --title "..." --method "Kuant" --sample "200" --findings "..." --gap "..." --relevance "Tinggi"

Usage import:
  from lit_review_matrix import create_matrix, add_entry
  create_matrix("matrix.csv")
  add_entry("matrix.csv", author="Smith", year=2023, ...)
"""

import argparse
import csv
import json
import os
from pathlib import Path

COLUMNS = [
    "No",
    "Author (Year)",
    "Title",
    "Method",
    "Sample Size",
    "Key Findings",
    "Research Gap",
    "Relevance to Study",
    "DOI/URL",
    "Notes"
]


def create_matrix(filepath: str):
    """Buat template matrix kosong."""
    ext = Path(filepath).suffix.lower()

    if ext == ".csv":
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(COLUMNS)
    elif ext == ".xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Literature Review"

            # Header styling
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col, header in enumerate(COLUMNS, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            # Set column widths
            widths = [5, 20, 35, 15, 12, 40, 30, 25, 20, 25]
            for col, w in enumerate(widths, 1):
                ws.column_dimensions[chr(64+col) if col <= 26 else 'A'].width = w

            # Freeze top row
            ws.freeze_panes = "A2"

            wb.save(filepath)
        except ImportError:
            # Fallback to CSV
            csv_path = filepath.replace('.xlsx', '.csv')
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(COLUMNS)
            print(f"openpyxl not installed, created CSV instead: {csv_path}")
            return csv_path
    else:
        # Default to CSV
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(COLUMNS)

    return filepath


def add_entry(filepath: str, **entry):
    """Tambah entry ke matrix."""
    # Read existing
    rows = []
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)

    # Determine row number
    if rows:
        last_no = 0
        for row in rows[1:]:
            if row and row[0].isdigit():
                last_no = int(row[0])
        entry_no = last_no + 1
    else:
        rows = [COLUMNS]
        entry_no = 1

    # Build row
    row = [
        entry_no,
        f"{entry.get('author', '')} ({entry.get('year', '')})",
        entry.get('title', ''),
        entry.get('method', ''),
        entry.get('sample', ''),
        entry.get('findings', ''),
        entry.get('gap', ''),
        entry.get('relevance', ''),
        entry.get('doi', '') or entry.get('url', ''),
        entry.get('notes', '')
    ]
    rows.append(row)

    # Write back
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    return entry_no


def import_from_bibtex(bibtex_file: str, output_csv: str):
    """Import entries dari BibTeX file ke CSV matrix."""
    import re

    with open(bibtex_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # Parse BibTeX entries
    entries = re.findall(r'@\w+\{([^,]+),(.*?)\n\}', content, re.DOTALL)

    create_matrix(output_csv)

    for key, body in entries:
        fields = {}
        for match in re.finditer(r'(\w+)\s*=\s*[{"](.*?)[}"]', body, re.DOTALL):
            fields[match.group(1).lower()] = match.group(2).strip()

        add_entry(output_csv,
            author=fields.get('author', ''),
            year=fields.get('year', ''),
            title=fields.get('title', ''),
            doi=fields.get('doi', ''),
            url=fields.get('url', '')
        )

    return len(entries)


def summary(filepath: str) -> dict:
    """Generate summary dari matrix."""
    if not os.path.exists(filepath):
        return {"error": "File not found"}

    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    total = len(rows)
    methods = {}
    gaps = []

    for row in rows:
        method = row.get("Method", "").strip()
        if method:
            methods[method] = methods.get(method, 0) + 1
        gap = row.get("Research Gap", "").strip()
        if gap:
            gaps.append(gap)

    return {
        "total_entries": total,
        "methods_breakdown": methods,
        "gaps_identified": len(gaps),
        "gaps": gaps[:10],
    }


def cli():
    parser = argparse.ArgumentParser(description="Literature Review Matrix Generator")
    sub = parser.add_subparsers(dest="action")

    # Create
    p_create = sub.add_parser("create", help="Create empty matrix")
    p_create.add_argument("--output", required=True, help="Output file (.csv or .xlsx)")

    # Add
    p_add = sub.add_parser("add", help="Add entry to matrix")
    p_add.add_argument("--file", required=True)
    p_add.add_argument("--author", default="")
    p_add.add_argument("--year", default="")
    p_add.add_argument("--title", default="")
    p_add.add_argument("--method", default="")
    p_add.add_argument("--sample", default="")
    p_add.add_argument("--findings", default="")
    p_add.add_argument("--gap", default="")
    p_add.add_argument("--relevance", default="")
    p_add.add_argument("--doi", default="")
    p_add.add_argument("--url", default="")
    p_add.add_argument("--notes", default="")

    # Import BibTeX
    p_import = sub.add_parser("import-bibtex", help="Import from BibTeX file")
    p_import.add_argument("--bibtex", required=True)
    p_import.add_argument("--output", required=True)

    # Summary
    p_sum = sub.add_parser("summary", help="Show matrix summary")
    p_sum.add_argument("--file", required=True)

    args = parser.parse_args()

    if args.action == "create":
        create_matrix(args.output)
        print(f"Created: {args.output}")

    elif args.action == "add":
        entry_no = add_entry(args.file,
            author=args.author, year=args.year, title=args.title,
            method=args.method, sample=args.sample, findings=args.findings,
            gap=args.gap, relevance=args.relevance, doi=args.doi,
            url=args.url, notes=args.notes)
        print(f"Added entry #{entry_no} to {args.file}")

    elif args.action == "import-bibtex":
        count = import_from_bibtex(args.bibtex, args.output)
        print(f"Imported {count} entries to {args.output}")

    elif args.action == "summary":
        result = summary(args.file)
        print(json.dumps(result, indent=2, ensure_ascii=False))

    else:
        parser.print_help()


if __name__ == "__main__":
    cli()
