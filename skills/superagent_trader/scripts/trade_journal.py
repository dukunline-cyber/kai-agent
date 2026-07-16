#!/usr/bin/env python3.12
"""
Superagent Trader — Trade Journal Generator
Creates Excel trade journal with auto formulas.

Usage:
    python trade_journal.py /tmp/Trade_Journal.xlsx
    python trade_journal.py /tmp/journal.xlsx --name "Teguh"
"""
import sys, argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


def create_journal(output_path: str, trader_name: str = "Trader"):
    wb = openpyxl.Workbook()
    
    # ─── Styles ───
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    title_font = Font(name="Arial", bold=True, size=16)
    subtitle_font = Font(name="Arial", size=10, color="666666")
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    center = Alignment(horizontal="center", vertical="center")
    
    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    # ═══ SHEET 1: TRADE LOG ═══
    ws = wb.active
    ws.title = "Trade Log"
    ws.sheet_properties.tabColor = "1a1a2e"
    
    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = f"📊 TRADE JOURNAL — {trader_name.upper()}"
    ws["A1"].font = title_font
    ws.merge_cells("A2:P2")
    ws["A2"] = f"Created: {datetime.now().strftime('%Y-%m-%d')} | Superagent Trader by Viktor AI"
    ws["A2"].font = subtitle_font
    ws.row_dimensions[1].height = 30
    
    # Headers (row 4)
    headers = [
        ("A", "No", 5),
        ("B", "Date", 12),
        ("C", "Session", 10),
        ("D", "Pair", 10),
        ("E", "Timeframe", 10),
        ("F", "Direction", 10),
        ("G", "Strategy", 15),
        ("H", "Entry", 12),
        ("I", "SL", 12),
        ("J", "TP", 12),
        ("K", "Lot Size", 10),
        ("L", "RR Ratio", 8),
        ("M", "Result", 10),
        ("N", "P&L ($)", 10),
        ("O", "P&L (pips)", 10),
        ("P", "Notes / Screenshot", 25),
    ]
    
    for col, title, width in headers:
        cell = ws[f"{col}4"]
        cell.value = title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[col].width = width
    
    # Data rows (5-104) with formulas
    for row in range(5, 105):
        for col_letter, _, _ in headers:
            cell = ws[f"{col_letter}{row}"]
            cell.border = border
            cell.alignment = center
        
        # No
        ws[f"A{row}"] = row - 4
        
        # RR Ratio formula: =IF(AND(H{row}<>"",I{row}<>"",J{row}<>""), ABS(J-H)/ABS(I-H), "")
        ws[f"L{row}"] = f'=IF(AND(H{row}<>"",I{row}<>"",J{row}<>""),ROUND(ABS(J{row}-H{row})/ABS(I{row}-H{row}),1),"")'
        
        # P&L pips formula
        ws[f"O{row}"] = f'=IF(AND(H{row}<>"",M{row}<>""),IF(M{row}="WIN",ABS(J{row}-H{row})*IF(OR(LEFT(D{row},3)="JPY",RIGHT(D{row},3)="JPY"),100,10000),IF(M{row}="LOSS",-ABS(I{row}-H{row})*IF(OR(LEFT(D{row},3)="JPY",RIGHT(D{row},3)="JPY"),100,10000),0)),"")'
    
    # Conditional formatting for Result column
    ws.conditional_formatting.add(f"M5:M104",
        CellIsRule(operator="equal", formula=['"WIN"'], fill=green_fill))
    ws.conditional_formatting.add(f"M5:M104",
        CellIsRule(operator="equal", formula=['"LOSS"'], fill=red_fill))
    
    # ─── Data validation for dropdowns ───
    from openpyxl.worksheet.datavalidation import DataValidation
    
    session_dv = DataValidation(type="list", formula1='"Sydney,Tokyo,London,New York,LDN-NY Overlap"')
    ws.add_data_validation(session_dv)
    session_dv.add(f"C5:C104")
    
    direction_dv = DataValidation(type="list", formula1='"BUY,SELL"')
    ws.add_data_validation(direction_dv)
    direction_dv.add(f"F5:F104")
    
    tf_dv = DataValidation(type="list", formula1='"M5,M15,M30,H1,H4,D,W"')
    ws.add_data_validation(tf_dv)
    tf_dv.add(f"E5:E104")
    
    result_dv = DataValidation(type="list", formula1='"WIN,LOSS,BE,OPEN"')
    ws.add_data_validation(result_dv)
    result_dv.add(f"M5:M104")
    
    strategy_dv = DataValidation(type="list", formula1='"SMC/ICT,EMA Cross,RSI OB/OS,MACD Cross,BB Bounce,Harmonic,Elliott Wave,Wyckoff,Price Action,Other"')
    ws.add_data_validation(strategy_dv)
    strategy_dv.add(f"G5:G104")
    
    # Freeze panes
    ws.freeze_panes = "A5"
    
    # ═══ SHEET 2: DASHBOARD ═══
    ds = wb.create_sheet("Dashboard")
    ds.sheet_properties.tabColor = "00c853"
    
    ds.merge_cells("A1:F1")
    ds["A1"] = "📈 PERFORMANCE DASHBOARD"
    ds["A1"].font = title_font
    
    metrics = [
        ("A3", "METRIC", header_font, header_fill),
        ("B3", "VALUE", header_font, header_fill),
        ("A4", "Total Trades"), ("B4", "=COUNTA('Trade Log'!M5:M104)-COUNTBLANK('Trade Log'!M5:M104)"),
        ("A5", "Wins"), ("B5", '=COUNTIF(\'Trade Log\'!M5:M104,"WIN")'),
        ("A6", "Losses"), ("B6", '=COUNTIF(\'Trade Log\'!M5:M104,"LOSS")'),
        ("A7", "Win Rate %"), ("B7", '=IF(B4>0,ROUND(B5/B4*100,1),0)'),
        ("A8", "Total P&L ($)"), ("B8", "=SUM('Trade Log'!N5:N104)"),
        ("A9", "Avg Win ($)"), ("B9", '=IF(B5>0,AVERAGEIF(\'Trade Log\'!M5:M104,"WIN",\'Trade Log\'!N5:N104),0)'),
        ("A10", "Avg Loss ($)"), ("B10", '=IF(B6>0,AVERAGEIF(\'Trade Log\'!M5:M104,"LOSS",\'Trade Log\'!N5:N104),0)'),
        ("A11", "Avg RR"), ("B11", '=IF(B4>0,ROUND(AVERAGE(\'Trade Log\'!L5:L104),2),0)'),
        ("A12", "Best Trade ($)"), ("B12", "=MAX('Trade Log'!N5:N104)"),
        ("A13", "Worst Trade ($)"), ("B13", "=MIN('Trade Log'!N5:N104)"),
        ("A14", "Total Pips"), ("B14", "=SUM('Trade Log'!O5:O104)"),
        ("A15", "Profit Factor"), ("B15", '=IF(ABS(SUMIF(\'Trade Log\'!M5:M104,"LOSS",\'Trade Log\'!N5:N104))>0,SUMIF(\'Trade Log\'!M5:M104,"WIN",\'Trade Log\'!N5:N104)/ABS(SUMIF(\'Trade Log\'!M5:M104,"LOSS",\'Trade Log\'!N5:N104)),0)'),
    ]
    
    for item in metrics:
        cell = ds[item[0]]
        cell.value = item[1]
        cell.border = border
        if len(item) > 2:
            cell.font = item[2]
            cell.fill = item[3]
            cell.alignment = center
    
    ds.column_dimensions["A"].width = 20
    ds.column_dimensions["B"].width = 15
    
    # Strategy breakdown
    ds["D3"] = "STRATEGY BREAKDOWN"
    ds["D3"].font = Font(name="Arial", bold=True, size=12)
    strategies = ["SMC/ICT", "EMA Cross", "RSI OB/OS", "MACD Cross", "BB Bounce",
                  "Harmonic", "Elliott Wave", "Wyckoff", "Price Action"]
    
    ds["D4"] = "Strategy"; ds["E4"] = "Trades"; ds["F4"] = "Wins"; ds["G4"] = "Win %"
    for c in ["D", "E", "F", "G"]:
        ds[f"{c}4"].font = header_font
        ds[f"{c}4"].fill = header_fill
        ds[f"{c}4"].alignment = center
        ds[f"{c}4"].border = border
    
    for i, strat in enumerate(strategies, 5):
        row = i
        ds[f"D{row}"] = strat
        ds[f"D{row}"].border = border
        ds[f"E{row}"] = f'=COUNTIF(\'Trade Log\'!G5:G104,D{row})'
        ds[f"E{row}"].border = border
        ds[f"F{row}"] = f'=COUNTIFS(\'Trade Log\'!G5:G104,D{row},\'Trade Log\'!M5:M104,"WIN")'
        ds[f"F{row}"].border = border
        ds[f"G{row}"] = f'=IF(E{row}>0,ROUND(F{row}/E{row}*100,1),0)'
        ds[f"G{row}"].border = border
    
    ds.column_dimensions["D"].width = 18
    ds.column_dimensions["E"].width = 10
    ds.column_dimensions["F"].width = 10
    ds.column_dimensions["G"].width = 10
    
    # ═══ SHEET 3: WEEKLY REVIEW ═══
    wr = wb.create_sheet("Weekly Review")
    wr.sheet_properties.tabColor = "2979FF"
    
    wr.merge_cells("A1:E1")
    wr["A1"] = "📅 WEEKLY REVIEW"
    wr["A1"].font = title_font
    
    review_headers = ["Week", "Trades", "Wins", "Losses", "P&L ($)", "Notes / Lessons Learned"]
    for i, h in enumerate(review_headers):
        col = get_column_letter(i + 1)
        cell = wr[f"{col}3"]
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        wr.column_dimensions[col].width = 15 if i < 5 else 35
    
    for row in range(4, 56):
        for col_i in range(1, 7):
            cell = wr[f"{get_column_letter(col_i)}{row}"]
            cell.border = border
    
    wr.freeze_panes = "A4"
    
    # ═══ SHEET 4: PRE-TRADE CHECKLIST ═══
    cl = wb.create_sheet("Checklist")
    cl.sheet_properties.tabColor = "FFD600"
    
    cl.merge_cells("A1:D1")
    cl["A1"] = "☑ PRE-TRADE CHECKLIST"
    cl["A1"].font = title_font
    
    checklist = [
        "1. Cek economic calendar — hindari high-impact news ±30 menit",
        "2. Analisa Key TF → tentukan bias (BULLISH / BEARISH / NEUTRAL)",
        "3. Identifikasi key levels di Key TF & Behavioral TF",
        "4. Cari setup di Behavioral TF yang sejalan dengan bias",
        "5. Tunggu konfirmasi di Entry TF (ChoCH/BOS + candlestick)",
        "6. Hitung position size & set SL/TP",
        "7. Cek RR ratio (minimum 1:2)",
        "8. Confluence check (minimum 3 faktor)",
        "9. Execute & screenshot untuk jurnal",
        "10. Set & forget — jangan ganggu trade yang sudah berjalan",
    ]
    
    for i, item in enumerate(checklist, 3):
        cl[f"A{i}"] = "☐"
        cl[f"A{i}"].alignment = center
        cl.merge_cells(f"B{i}:D{i}")
        cl[f"B{i}"] = item
        cl[f"B{i}"].font = Font(name="Arial", size=11)
    
    cl.column_dimensions["A"].width = 5
    cl.column_dimensions["B"].width = 60
    
    # Save
    wb.save(output_path)
    print(f"Trade journal saved: {output_path}")
    print(f"Sheets: Trade Log, Dashboard, Weekly Review, Checklist")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Trade Journal Generator")
    parser.add_argument("output", type=str, nargs="?", default="/tmp/Trade_Journal.xlsx",
                       help="Output Excel file path")
    parser.add_argument("--name", type=str, default="Trader", help="Trader name")
    
    args = parser.parse_args()
    create_journal(args.output, args.name)


if __name__ == "__main__":
    main()
